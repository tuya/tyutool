#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import shutil
import sys
from datetime import datetime
from openpyxl import load_workbook

# Accepted header aliases (lowercase) -> canonical role
_HEADER_ALIASES = {
    "uuid": "uuid",
    "authkey": "authkey",
    "key": "authkey",
    "status": "status",
    "mac": "mac",
    "timestamp": "timestamp",
}


class AuthExcelParser:
    """Read and write TuyaOpen authorization Excel files (.xlsx).

    Locates UUID / AUTHKEY columns by header name (case-insensitive),
    accepting common aliases such as 'key' for AUTHKEY.
    STATUS, MAC, TIMESTAMP columns are auto-created when first needed.
    """

    STATUS_USED = "USED"

    def __init__(self):
        self.filepath = None
        self.wb = None
        self.ws = None
        self._lock_fd = None
        self._backed_up = False
        self._col = {}  # role -> column number

    def load(self, filepath):
        """Load an Excel file and parse authorization codes."""
        self.filepath = filepath
        self._backed_up = False
        self.wb = load_workbook(filepath)
        self.ws = self.wb.active
        self._detect_columns()

    def _detect_columns(self):
        """Scan header row and map column roles by name aliases."""
        self._col = {}
        for col in range(1, self.ws.max_column + 1):
            val = self.ws.cell(row=1, column=col).value
            if val is None:
                continue
            role = _HEADER_ALIASES.get(str(val).strip().lower())
            if role and role not in self._col:
                self._col[role] = col

        if "uuid" not in self._col or "authkey" not in self._col:
            raise ValueError(
                "Missing required columns in Excel header: UUID and AUTHKEY (or key) columns are needed"
            )

        next_col = self.ws.max_column + 1
        for role in ("status", "mac", "timestamp"):
            if role not in self._col:
                self._col[role] = next_col
                next_col += 1

    def _backup(self):
        """Create a .bak copy of the Excel file if not already backed up."""
        if self._backed_up:
            return
        bak_path = self.filepath + ".bak"
        if not os.path.exists(bak_path):
            shutil.copy2(self.filepath, bak_path)
        self._backed_up = True

    def _ensure_header(self):
        canonical = {
            "uuid": "UUID", "authkey": "AUTHKEY",
            "status": "STATUS", "mac": "MAC", "timestamp": "TIMESTAMP",
        }
        for role, col in self._col.items():
            cell = self.ws.cell(row=1, column=col)
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = canonical.get(role, role.upper())

    def get_stats(self):
        """Return (total, used, remain) counts."""
        total = 0
        used = 0
        col_uuid = self._col["uuid"]
        col_status = self._col["status"]
        for row in range(2, self.ws.max_row + 1):
            uuid_val = self.ws.cell(row=row, column=col_uuid).value
            if not uuid_val:
                continue
            total += 1
            status = self.ws.cell(row=row, column=col_status).value
            if status and str(status).strip().upper() == self.STATUS_USED:
                used += 1
        return (total, used, total - used)

    def get_next_available(self):
        """Return (row_number, uuid, authkey) of the next unused code, or None."""
        col_uuid = self._col["uuid"]
        col_authkey = self._col["authkey"]
        col_status = self._col["status"]
        for row in range(2, self.ws.max_row + 1):
            uuid_val = self.ws.cell(row=row, column=col_uuid).value
            authkey_val = self.ws.cell(row=row, column=col_authkey).value
            if not uuid_val or not authkey_val:
                continue
            status = self.ws.cell(row=row, column=col_status).value
            if status and str(status).strip().upper() == self.STATUS_USED:
                continue
            return (row, str(uuid_val).strip(), str(authkey_val).strip())
        return None

    def find_by_uuid(self, uuid):
        """Return row number for the given UUID, or None if not found."""
        col_uuid = self._col["uuid"]
        for row in range(2, self.ws.max_row + 1):
            val = self.ws.cell(row=row, column=col_uuid).value
            if val and str(val).strip() == uuid:
                return row
        return None

    def find_by_mac(self, mac):
        """Return (row, uuid, authkey) for the given MAC, or None if not found."""
        col_mac = self._col.get("mac")
        if col_mac is None:
            return None
        col_uuid = self._col["uuid"]
        col_authkey = self._col["authkey"]
        for row in range(2, self.ws.max_row + 1):
            val = self.ws.cell(row=row, column=col_mac).value
            if val and str(val).strip().upper() == mac.upper():
                uuid_val = str(self.ws.cell(row=row, column=col_uuid).value or "").strip()
                authkey_val = str(self.ws.cell(row=row, column=col_authkey).value or "").strip()
                return (row, uuid_val, authkey_val)
        return None

    def mark_used(self, row, mac, timestamp=None):
        """Mark a row as USED with MAC and timestamp, then save."""
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._backup()
        self._ensure_header()
        self.ws.cell(row=row, column=self._col["status"], value=self.STATUS_USED)
        self.ws.cell(row=row, column=self._col["mac"], value=mac)
        self.ws.cell(row=row, column=self._col["timestamp"], value=timestamp)
        self.wb.save(self.filepath)

    def lock(self):
        """Acquire an advisory file lock to prevent concurrent writes."""
        lock_path = self.filepath + ".lock"
        self._lock_fd = open(lock_path, 'w')
        try:
            if sys.platform == 'win32':
                import msvcrt
                msvcrt.locking(self._lock_fd.fileno(), msvcrt.LK_NBLCK, 1)
            else:
                import fcntl
                fcntl.flock(self._lock_fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
            return True
        except (IOError, OSError):
            self._lock_fd.close()
            self._lock_fd = None
            return False

    def unlock(self):
        """Release the advisory file lock."""
        if self._lock_fd:
            try:
                if sys.platform == 'win32':
                    import msvcrt
                    msvcrt.locking(self._lock_fd.fileno(), msvcrt.LK_UNLCK, 1)
                else:
                    import fcntl
                    fcntl.flock(self._lock_fd, fcntl.LOCK_UN)
                self._lock_fd.close()
            except Exception:
                pass
            self._lock_fd = None
            lock_path = self.filepath + ".lock"
            try:
                os.remove(lock_path)
            except OSError:
                pass

    def close(self):
        """Close workbook and release lock."""
        self.unlock()
        if self.wb:
            try:
                self.wb.close()
            except Exception:
                pass
            self.wb = None
            self.ws = None
